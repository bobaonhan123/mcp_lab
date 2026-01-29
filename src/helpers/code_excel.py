"""Code to Excel helper for writing code blocks to Excel files with syntax highlighting."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .code import CodeBlock, capture_code_block, capture_multiple_blocks, detect_language
from .plantuml import generate_plantuml_image

try:
    from openpyxl.drawing.image import Image as XLImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

# Syntax highlighting with Pygments
try:
    from pygments import lex
    from pygments.lexers import get_lexer_by_name, guess_lexer
    from pygments.token import Token
    HAS_PYGMENTS = True
except ImportError:
    HAS_PYGMENTS = False


# Monokai-inspired color scheme for syntax highlighting
SYNTAX_COLORS = {
    Token.Keyword: "F92672",           # Pink/Magenta - keywords (def, class, if, etc.)
    Token.Keyword.Namespace: "F92672",
    Token.Keyword.Declaration: "66D9EF",  # Cyan - type declarations
    Token.Name.Function: "A6E22E",     # Green - function names
    Token.Name.Class: "A6E22E",        # Green - class names
    Token.Name.Decorator: "A6E22E",    # Green - decorators
    Token.Name.Builtin: "66D9EF",      # Cyan - builtins (print, len, etc.)
    Token.String: "E6DB74",            # Yellow - strings
    Token.String.Doc: "75715E",        # Gray - docstrings
    Token.Comment: "75715E",           # Gray - comments
    Token.Comment.Single: "75715E",
    Token.Number: "AE81FF",            # Purple - numbers
    Token.Operator: "F92672",          # Pink - operators
    Token.Punctuation: "F8F8F2",       # White - punctuation
    Token.Name: "F8F8F2",              # White - names
}


def _get_token_color(token_type) -> str | None:
    """Get the color for a token type, checking parent types."""
    while token_type:
        if token_type in SYNTAX_COLORS:
            return SYNTAX_COLORS[token_type]
        token_type = token_type.parent
    return None


def _get_lexer_for_language(language: str):
    """Get Pygments lexer for a language."""
    try:
        return get_lexer_by_name(language)
    except:
        return get_lexer_by_name("python")  # Default to Python


def _apply_code_style(ws, start_row: int, code_block: CodeBlock, use_dark_theme: bool = True) -> int:
    """Apply styling and write code block to worksheet. Returns next available row."""
    # Theme colors
    if use_dark_theme:
        bg_color = "272822"  # Monokai dark background
        header_bg = "1E1F1C"
        line_num_color = "75715E"
        default_text_color = "F8F8F2"
    else:
        bg_color = "FAFAFA"
        header_bg = "366092"
        line_num_color = "888888"
        default_text_color = "333333"
    
    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
    bg_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    code_font = Font(name="Consolas", size=10, color=default_text_color)
    line_num_font = Font(name="Consolas", size=10, color=line_num_color)
    border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )
    
    # Write file header
    header_cell = ws.cell(row=start_row, column=1)
    header_cell.value = f"📄 {code_block.file_path} (Lines {code_block.start_line}-{code_block.end_line})"
    header_cell.font = header_font
    header_cell.fill = header_fill
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    
    # Column headers
    col_header_font = Font(bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    
    line_header = ws.cell(row=start_row + 1, column=1, value="Line")
    line_header.font = col_header_font
    line_header.fill = col_header_fill
    line_header.alignment = Alignment(horizontal="center")
    
    code_header = ws.cell(row=start_row + 1, column=2, value="Code")
    code_header.font = col_header_font
    code_header.fill = col_header_fill
    code_header.alignment = Alignment(horizontal="left")
    
    # Get lexer for syntax highlighting
    lexer = None
    if HAS_PYGMENTS:
        lang = detect_language(code_block.file_path)
        lexer = _get_lexer_for_language(lang)
    
    # Write code lines
    lines = code_block.code.rstrip("\n").split("\n")
    current_row = start_row + 2
    
    for i, line in enumerate(lines, start=code_block.start_line):
        # Line number cell
        line_cell = ws.cell(row=current_row, column=1, value=i)
        line_cell.font = line_num_font
        line_cell.fill = bg_fill
        line_cell.alignment = Alignment(horizontal="right")
        line_cell.border = border
        
        # Code cell - apply syntax highlighting color for dominant token
        code_cell = ws.cell(row=current_row, column=2, value=line)
        code_cell.fill = bg_fill
        code_cell.alignment = Alignment(horizontal="left")
        code_cell.border = border
        
        # Determine line color based on content
        font_color = default_text_color
        if HAS_PYGMENTS and lexer and line.strip():
            # Get tokens for this line
            tokens = list(lex(line + "\n", lexer))
            # Find the most significant token for coloring the line
            for token_type, token_value in tokens:
                if token_value.strip():  # Skip whitespace
                    color = _get_token_color(token_type)
                    if color:
                        font_color = color
                        break
        
        code_cell.font = Font(name="Consolas", size=10, color=font_color)
        
        current_row += 1
    
    return current_row + 1  # Leave one blank row


def write_code_blocks_to_excel(
    code_blocks: list[CodeBlock] | CodeBlock,
    output_path: str | Path,
    sheet_name: str = "Code Blocks",
    title: str = "Code Documentation",
) -> Path:
    """Write code blocks to a new Excel file.
    
    Args:
        code_blocks: Single CodeBlock or list of CodeBlocks
        output_path: Path for the output Excel file
        sheet_name: Name of the worksheet
        title: Title for the document
        
    Returns:
        Path to the saved file
    """
    if isinstance(code_blocks, CodeBlock):
        code_blocks = [code_blocks]
    
    output = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Set column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 120
    
    # Title
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    
    current_row = 3
    for block in code_blocks:
        current_row = _apply_code_style(ws, current_row, block)
    
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def write_code_and_diagram_to_excel(
    code_blocks: list[CodeBlock] | CodeBlock,
    puml_code: str,
    output_path: str | Path,
    server_url: str = "http://localhost:8080",
    code_sheet_name: str = "Code Blocks",
    diagram_sheet_name: str = "Sequence Diagram",
) -> Path:
    """Write code blocks and PlantUML diagram to Excel with separate sheets.
    
    Args:
        code_blocks: Single CodeBlock or list of CodeBlocks
        puml_code: PlantUML diagram code
        output_path: Path for the output Excel file
        server_url: PlantUML server URL
        code_sheet_name: Name of the code blocks sheet
        diagram_sheet_name: Name of the diagram sheet
        
    Returns:
        Path to the saved file
    """
    if isinstance(code_blocks, CodeBlock):
        code_blocks = [code_blocks]
    
    output = Path(output_path)
    wb = Workbook()
    
    # =========================================================================
    # Sheet 1: Code Blocks
    # =========================================================================
    ws_code = wb.active
    ws_code.title = code_sheet_name
    
    # Set column widths
    ws_code.column_dimensions["A"].width = 8
    ws_code.column_dimensions["B"].width = 120
    
    # Title
    title_cell = ws_code.cell(row=1, column=1, value="Code Documentation")
    title_cell.font = Font(bold=True, size=14)
    ws_code.merge_cells("A1:B1")
    
    current_row = 3
    for block in code_blocks:
        current_row = _apply_code_style(ws_code, current_row, block)
    
    # =========================================================================
    # Sheet 2: Sequence Diagram
    # =========================================================================
    ws_diagram = wb.create_sheet(title=diagram_sheet_name)
    
    # Set column widths
    ws_diagram.column_dimensions["A"].width = 15
    ws_diagram.column_dimensions["B"].width = 100
    
    # Title
    ws_diagram.cell(row=1, column=1, value="Sequence Diagram").font = Font(bold=True, size=14)
    ws_diagram.merge_cells("A1:B1")
    
    # PlantUML code section
    ws_diagram.cell(row=3, column=1, value="PlantUML Code:").font = Font(bold=True)
    code_cell = ws_diagram.cell(row=4, column=1, value=puml_code)
    code_cell.font = Font(name="Consolas", size=10)
    code_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_diagram.row_dimensions[4].height = 200
    ws_diagram.merge_cells("A4:B4")
    
    # Generate and add diagram image
    if HAS_PILLOW:
        try:
            image_data = generate_plantuml_image(puml_code, server_url=server_url)
            img_stream = BytesIO(image_data)
            img = XLImage(img_stream)
            
            # Place image
            ws_diagram.cell(row=6, column=1, value="Generated Diagram:").font = Font(bold=True)
            ws_diagram.add_image(img, "A7")
        except Exception as e:
            ws_diagram.cell(row=6, column=1, value=f"Error generating diagram: {e}")
    else:
        ws_diagram.cell(row=6, column=1, value="(Pillow not installed - cannot embed images)")
    
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def capture_and_write_to_excel(
    file_path: str | Path,
    ranges: list[tuple[int, int | None]],
    output_path: str | Path,
    title: str = "Code Documentation",
) -> Path:
    """Capture code blocks from a file and write to Excel in one step.
    
    Args:
        file_path: Source code file path
        ranges: List of (start_line, end_line) tuples
        output_path: Output Excel file path
        title: Document title
        
    Returns:
        Path to the saved file
    """
    blocks = capture_multiple_blocks(file_path, ranges)
    return write_code_blocks_to_excel(blocks, output_path, title=title)
