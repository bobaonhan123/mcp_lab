"""Combined Excel helper for creating multi-sheet documentation files."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .code import CodeBlock, capture_multiple_blocks, detect_language
from .diff import compute_diff, DiffLine
from .search import SearchSummary, search_in_folder
from .plantuml import generate_plantuml_image

try:
    from openpyxl.drawing.image import Image as XLImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

try:
    from pygments import lex
    from pygments.lexers import get_lexer_by_name
    from pygments.token import Token
    HAS_PYGMENTS = True
except ImportError:
    HAS_PYGMENTS = False


# Monokai color scheme
COLORS = {
    "bg": "272822",
    "header_bg": "1E1F1C",
    "col_header_bg": "444444",
    "text": "F8F8F2",
    "line_num": "75715E",
    "added_bg": "1E3A1E",
    "added_text": "A6E22E",
    "removed_bg": "3A1E1E",
    "removed_text": "F92672",
    "match_bg": "3A3A1E",
    "match_text": "E6DB74",
    "context_bg": "272822",
    "context_text": "75715E",
    "summary_bg": "366092",
}


def _get_border():
    return Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )


def _write_code_sheet(wb: Workbook, code_blocks: list[CodeBlock], sheet_name: str = "Code"):
    """Write code blocks to a sheet."""
    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 120
    
    border = _get_border()
    bg_fill = PatternFill(start_color=COLORS["bg"], end_color=COLORS["bg"], fill_type="solid")
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    col_header_fill = PatternFill(start_color=COLORS["col_header_bg"], end_color=COLORS["col_header_bg"], fill_type="solid")
    
    row = 1
    for block in code_blocks:
        # File header
        header = ws.cell(row=row, column=1, value=f"📄 {block.file_path} (Lines {block.start_line}-{block.end_line})")
        header.font = Font(bold=True, size=11, color="FFFFFF")
        header.fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        
        # Column headers
        for col, text in [(1, "Line"), (2, "Code")]:
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = col_header_fill
        row += 1
        
        # Code lines
        lines = block.code.rstrip("\n").split("\n")
        for i, line in enumerate(lines, start=block.start_line):
            line_cell = ws.cell(row=row, column=1, value=i)
            line_cell.font = Font(name="Consolas", size=10, color=COLORS["line_num"])
            line_cell.fill = bg_fill
            line_cell.alignment = Alignment(horizontal="right")
            line_cell.border = border
            
            code_cell = ws.cell(row=row, column=2, value=line)
            code_cell.font = Font(name="Consolas", size=10, color=COLORS["text"])
            code_cell.fill = bg_fill
            code_cell.border = border
            row += 1
        
        row += 1  # Blank row between blocks


def _write_diff_sheet(wb: Workbook, old_code: str, new_code: str, 
                       sheet_name: str = "Diff", file_path: str | None = None,
                       old_label: str = "Before", new_label: str = "After"):
    """Write GitLab-style side-by-side diff view to a sheet."""
    ws = wb.create_sheet(title=sheet_name)
    
    # Column widths for side-by-side view
    ws.column_dimensions["A"].width = 6   # Old line num
    ws.column_dimensions["B"].width = 60  # Old code
    ws.column_dimensions["C"].width = 3   # Separator
    ws.column_dimensions["D"].width = 6   # New line num
    ws.column_dimensions["E"].width = 60  # New code
    
    border = _get_border()
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    
    # Colors for diff - GitLab style
    REMOVED_BG = "FFEEF0"      # Light red background
    REMOVED_TEXT = "B31D28"    # Dark red text
    ADDED_BG = "E6FFED"        # Light green background  
    ADDED_TEXT = "22863A"      # Dark green text
    UNCHANGED_BG = "FFFFFF"    # White
    UNCHANGED_TEXT = "24292E"  # Dark gray
    LINE_NUM_BG = "F6F8FA"     # Light gray for line numbers
    LINE_NUM_TEXT = "6A737D"   # Gray text
    SEPARATOR_BG = "E1E4E8"    # Separator column
    
    # Header with file path
    header_text = f"📊 {file_path}" if file_path else "📊 Diff View"
    header = ws.cell(row=1, column=1, value=header_text)
    header.font = Font(bold=True, size=12, color="FFFFFF")
    header.fill = header_fill
    ws.merge_cells("A1:E1")
    
    # Column headers - Before | After
    col_header_fill = PatternFill(start_color="586069", end_color="586069", fill_type="solid")
    
    # Before header
    before_header = ws.cell(row=2, column=1, value=f"🔴 {old_label}")
    before_header.font = Font(bold=True, size=11, color="FFFFFF")
    before_header.fill = PatternFill(start_color="CB2431", end_color="CB2431", fill_type="solid")
    ws.merge_cells("A2:B2")
    
    # Separator
    sep_cell = ws.cell(row=2, column=3, value="")
    sep_cell.fill = PatternFill(start_color=SEPARATOR_BG, end_color=SEPARATOR_BG, fill_type="solid")
    
    # After header
    after_header = ws.cell(row=2, column=4, value=f"🟢 {new_label}")
    after_header.font = Font(bold=True, size=11, color="FFFFFF")
    after_header.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    ws.merge_cells("D2:E2")
    
    # Sub headers
    sub_header_fill = PatternFill(start_color="FAFBFC", end_color="FAFBFC", fill_type="solid")
    for col, text in [(1, "#"), (2, "Code"), (4, "#"), (5, "Code")]:
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = Font(bold=True, size=9, color="586069")
        cell.fill = sub_header_fill
        cell.border = border
    ws.cell(row=3, column=3).fill = PatternFill(start_color=SEPARATOR_BG, end_color=SEPARATOR_BG, fill_type="solid")
    
    # Compute diff
    diff_lines = compute_diff(old_code, new_code)
    
    # Build side-by-side view
    row = 4
    
    # Group consecutive changes for better visualization
    i = 0
    while i < len(diff_lines):
        diff_line = diff_lines[i]
        
        if diff_line.change_type == "unchanged":
            # Unchanged line - show on both sides
            _write_diff_row(ws, row, 
                           diff_line.line_num_old, diff_line.content,
                           diff_line.line_num_new, diff_line.content,
                           UNCHANGED_BG, UNCHANGED_TEXT, LINE_NUM_BG, LINE_NUM_TEXT, 
                           SEPARATOR_BG, border)
            row += 1
            i += 1
            
        elif diff_line.change_type == "removed":
            # Collect consecutive removed lines
            removed_lines = []
            while i < len(diff_lines) and diff_lines[i].change_type == "removed":
                removed_lines.append(diff_lines[i])
                i += 1
            
            # Collect consecutive added lines
            added_lines = []
            while i < len(diff_lines) and diff_lines[i].change_type == "added":
                added_lines.append(diff_lines[i])
                i += 1
            
            # Write paired lines
            max_lines = max(len(removed_lines), len(added_lines))
            for j in range(max_lines):
                old_num = removed_lines[j].line_num_old if j < len(removed_lines) else None
                old_content = removed_lines[j].content if j < len(removed_lines) else ""
                old_bg = REMOVED_BG if j < len(removed_lines) else UNCHANGED_BG
                old_text = REMOVED_TEXT if j < len(removed_lines) else UNCHANGED_TEXT
                
                new_num = added_lines[j].line_num_new if j < len(added_lines) else None
                new_content = added_lines[j].content if j < len(added_lines) else ""
                new_bg = ADDED_BG if j < len(added_lines) else UNCHANGED_BG
                new_text = ADDED_TEXT if j < len(added_lines) else UNCHANGED_TEXT
                
                _write_diff_row_split(ws, row,
                                      old_num, old_content, old_bg, old_text,
                                      new_num, new_content, new_bg, new_text,
                                      LINE_NUM_BG, LINE_NUM_TEXT, SEPARATOR_BG, border)
                row += 1
                
        elif diff_line.change_type == "added":
            # Added only (no corresponding removed)
            _write_diff_row_split(ws, row,
                                  None, "", UNCHANGED_BG, UNCHANGED_TEXT,
                                  diff_line.line_num_new, diff_line.content, ADDED_BG, ADDED_TEXT,
                                  LINE_NUM_BG, LINE_NUM_TEXT, SEPARATOR_BG, border)
            row += 1
            i += 1
        else:
            i += 1


def _write_diff_row(ws, row: int, 
                    old_num, old_content: str,
                    new_num, new_content: str,
                    bg: str, text: str, line_bg: str, line_text: str,
                    sep_bg: str, border):
    """Write a single diff row (same content on both sides)."""
    bg_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    line_fill = PatternFill(start_color=line_bg, end_color=line_bg, fill_type="solid")
    sep_fill = PatternFill(start_color=sep_bg, end_color=sep_bg, fill_type="solid")
    
    # Old side
    cell = ws.cell(row=row, column=1, value=old_num or "")
    cell.font = Font(name="Consolas", size=9, color=line_text)
    cell.fill = line_fill
    cell.alignment = Alignment(horizontal="right")
    cell.border = border
    
    cell = ws.cell(row=row, column=2, value=old_content)
    cell.font = Font(name="Consolas", size=10, color=text)
    cell.fill = bg_fill
    cell.border = border
    
    # Separator
    ws.cell(row=row, column=3).fill = sep_fill
    
    # New side
    cell = ws.cell(row=row, column=4, value=new_num or "")
    cell.font = Font(name="Consolas", size=9, color=line_text)
    cell.fill = line_fill
    cell.alignment = Alignment(horizontal="right")
    cell.border = border
    
    cell = ws.cell(row=row, column=5, value=new_content)
    cell.font = Font(name="Consolas", size=10, color=text)
    cell.fill = bg_fill
    cell.border = border


def _write_diff_row_split(ws, row: int,
                          old_num, old_content: str, old_bg: str, old_text: str,
                          new_num, new_content: str, new_bg: str, new_text: str,
                          line_bg: str, line_text: str, sep_bg: str, border):
    """Write a diff row with different styling for each side."""
    old_bg_fill = PatternFill(start_color=old_bg, end_color=old_bg, fill_type="solid")
    new_bg_fill = PatternFill(start_color=new_bg, end_color=new_bg, fill_type="solid")
    line_fill = PatternFill(start_color=line_bg, end_color=line_bg, fill_type="solid")
    sep_fill = PatternFill(start_color=sep_bg, end_color=sep_bg, fill_type="solid")
    
    # Old side
    cell = ws.cell(row=row, column=1, value=old_num if old_num else "")
    cell.font = Font(name="Consolas", size=9, color=line_text if old_num else old_text)
    cell.fill = line_fill if old_num else old_bg_fill
    cell.alignment = Alignment(horizontal="right")
    cell.border = border
    
    cell = ws.cell(row=row, column=2, value=old_content)
    cell.font = Font(name="Consolas", size=10, color=old_text)
    cell.fill = old_bg_fill
    cell.border = border
    
    # Separator
    ws.cell(row=row, column=3).fill = sep_fill
    
    # New side
    cell = ws.cell(row=row, column=4, value=new_num if new_num else "")
    cell.font = Font(name="Consolas", size=9, color=line_text if new_num else new_text)
    cell.fill = line_fill if new_num else new_bg_fill
    cell.alignment = Alignment(horizontal="right")
    cell.border = border
    
    cell = ws.cell(row=row, column=5, value=new_content)
    cell.font = Font(name="Consolas", size=10, color=new_text)
    cell.fill = new_bg_fill
    cell.border = border


def _write_search_sheet(wb: Workbook, summary: SearchSummary, sheet_name: str = "Search"):
    """Write search results to a sheet."""
    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 100
    
    border = _get_border()
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    col_header_fill = PatternFill(start_color=COLORS["col_header_bg"], end_color=COLORS["col_header_bg"], fill_type="solid")
    match_fill = PatternFill(start_color=COLORS["match_bg"], end_color=COLORS["match_bg"], fill_type="solid")
    context_fill = PatternFill(start_color=COLORS["context_bg"], end_color=COLORS["context_bg"], fill_type="solid")
    
    # Header
    header = ws.cell(row=1, column=1, value=f"🔍 Search: '{summary.query}' ({summary.total_matches} matches in {summary.files_with_matches} files)")
    header.font = Font(bold=True, size=11, color="FFFFFF")
    header.fill = header_fill
    ws.merge_cells("A1:C1")
    
    # Column headers
    for col, text in [(1, "File"), (2, "Line"), (3, "Content")]:
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = col_header_fill
    
    row = 3
    for result in summary.results:
        for match in result.matches:
            # Context before
            for ctx_i, ctx_line in enumerate(match.context_before):
                ctx_line_num = match.line_number - len(match.context_before) + ctx_i
                ws.cell(row=row, column=1, value="").fill = context_fill
                
                cell = ws.cell(row=row, column=2, value=ctx_line_num)
                cell.font = Font(name="Consolas", size=9, color=COLORS["context_text"])
                cell.fill = context_fill
                cell.border = border
                
                cell = ws.cell(row=row, column=3, value=ctx_line)
                cell.font = Font(name="Consolas", size=9, color=COLORS["context_text"])
                cell.fill = context_fill
                cell.border = border
                row += 1
            
            # Match line
            cell = ws.cell(row=row, column=1, value=match.file_path)
            cell.font = Font(size=9, color=COLORS["text"])
            cell.fill = match_fill
            cell.border = border
            
            cell = ws.cell(row=row, column=2, value=match.line_number)
            cell.font = Font(name="Consolas", size=10, bold=True, color=COLORS["match_text"])
            cell.fill = match_fill
            cell.border = border
            
            cell = ws.cell(row=row, column=3, value=match.line_content)
            cell.font = Font(name="Consolas", size=10, color=COLORS["match_text"])
            cell.fill = match_fill
            cell.border = border
            row += 1
            
            # Context after
            for ctx_i, ctx_line in enumerate(match.context_after):
                ctx_line_num = match.line_number + ctx_i + 1
                ws.cell(row=row, column=1, value="").fill = context_fill
                
                cell = ws.cell(row=row, column=2, value=ctx_line_num)
                cell.font = Font(name="Consolas", size=9, color=COLORS["context_text"])
                cell.fill = context_fill
                cell.border = border
                
                cell = ws.cell(row=row, column=3, value=ctx_line)
                cell.font = Font(name="Consolas", size=9, color=COLORS["context_text"])
                cell.fill = context_fill
                cell.border = border
                row += 1
            
            row += 1  # Blank row


def write_combined_excel(
    output_path: str | Path,
    code_blocks: list[CodeBlock] | None = None,
    diff_old: str | None = None,
    diff_new: str | None = None,
    diff_file_path: str | None = None,
    search_summary: SearchSummary | None = None,
    puml_code: str | None = None,
    puml_server_url: str = "http://localhost:8080",
) -> Path:
    """Create a combined Excel file with multiple sheets.
    
    Args:
        output_path: Path for output Excel file
        code_blocks: List of code blocks for "Code" sheet
        diff_old: Old code for diff comparison
        diff_new: New code for diff comparison
        diff_file_path: Optional file path for diff header
        search_summary: Search results for "Search" sheet
        puml_code: PlantUML code for "Diagram" sheet
        puml_server_url: PlantUML server URL
        
    Returns:
        Path to saved file
    """
    output = Path(output_path)
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Add requested sheets
    if code_blocks:
        _write_code_sheet(wb, code_blocks, "Code")
    
    if diff_old is not None and diff_new is not None:
        _write_diff_sheet(wb, diff_old, diff_new, "Diff", diff_file_path)
    
    if search_summary:
        _write_search_sheet(wb, search_summary, "Search")
    
    if puml_code and HAS_PILLOW:
        ws = wb.create_sheet(title="Diagram")
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 100
        
        header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        ws.cell(row=1, column=1, value="📊 Diagram").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=1, column=1).fill = header_fill
        ws.merge_cells("A1:B1")
        
        # PlantUML code
        ws.cell(row=3, column=1, value="PlantUML:").font = Font(bold=True)
        code_cell = ws.cell(row=4, column=1, value=puml_code)
        code_cell.font = Font(name="Consolas", size=10)
        code_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[4].height = 150
        ws.merge_cells("A4:B4")
        
        try:
            image_data = generate_plantuml_image(puml_code, server_url=puml_server_url)
            img_stream = BytesIO(image_data)
            img = XLImage(img_stream)
            ws.cell(row=6, column=1, value="Generated:").font = Font(bold=True)
            ws.add_image(img, "A7")
        except Exception as e:
            ws.cell(row=6, column=1, value=f"Error: {e}")
    
    # Remove empty default sheet if we have other sheets
    if len(wb.sheetnames) > 1 and default_sheet.title in wb.sheetnames:
        wb.remove(default_sheet)
    
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output
