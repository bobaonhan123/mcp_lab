"""PlantUML helper for generating diagrams and embedding them in Excel."""

from __future__ import annotations

import base64
import zlib
from io import BytesIO
from pathlib import Path
from typing import Literal

import httpx
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet

from .excel import MarkerNotFoundError


# PlantUML encoding alphabet (similar to base64 but different mapping)
PLANTUML_ALPHABET = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz-_"
BASE64_ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/"


def _encode_plantuml(text: str) -> str:
    """Encode PlantUML text to the format expected by PlantUML server.
    
    PlantUML uses a custom encoding: compress with deflate, then encode with
    a modified base64 alphabet.
    """
    # Compress with zlib (deflate)
    compressed = zlib.compress(text.encode("utf-8"), 9)[2:-4]  # Strip zlib header/trailer
    
    # Encode to base64
    b64 = base64.b64encode(compressed).decode("ascii")
    
    # Translate from standard base64 to PlantUML alphabet
    translation = str.maketrans(BASE64_ALPHABET, PLANTUML_ALPHABET)
    return b64.translate(translation)


def generate_plantuml_image(
    puml_code: str,
    server_url: str = "http://localhost:8080",
    output_format: Literal["png", "svg", "txt"] = "png",
    timeout: float = 30.0,
) -> bytes:
    """Generate a PlantUML diagram image by calling the PlantUML server.
    
    Args:
        puml_code: The PlantUML code (with or without @startuml/@enduml)
        server_url: Base URL of the PlantUML server (e.g., http://localhost:8080)
        output_format: Output format - 'png', 'svg', or 'txt'
        timeout: Request timeout in seconds
        
    Returns:
        The image data as bytes
        
    Raises:
        httpx.HTTPStatusError: If the server returns an error
        httpx.RequestError: If the request fails
    """
    # Ensure the code has @startuml/@enduml wrapper
    code = puml_code.strip()
    if not code.startswith("@start"):
        code = f"@startuml\n{code}\n@enduml"
    
    # Encode the PlantUML code
    encoded = _encode_plantuml(code)
    
    # Build the URL
    url = f"{server_url.rstrip('/')}/{output_format}/{encoded}"
    
    # Make the request
    with httpx.Client(timeout=timeout) as client:
        response = client.get(url)
        response.raise_for_status()
        return response.content


def _find_marker_cell(ws: Worksheet, marker: str):
    """Locate the cell whose value matches the marker string."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == marker:
                return cell
    raise MarkerNotFoundError(f"Could not find marker cell with value '{marker}'.")


PLANTUML_MARKER = "marker[plantuml]"
PLANTUML_IMAGE_MARKER = "marker[plantuml_image]"


def write_plantuml_to_excel(
    template_path: str | Path,
    output_path: str | Path,
    puml_code: str,
    server_url: str = "http://localhost:8080",
    image_width: int | None = None,
    image_height: int | None = None,
) -> Path:
    """Write PlantUML code and its rendered image to an Excel file.
    
    The function uses marker cells in the workbook:
    - marker[plantuml]: The PlantUML source code will be written below this cell
    - marker[plantuml_image]: The generated image will be placed at this cell
    
    Args:
        template_path: Path to the Excel template file
        output_path: Path where the output file will be saved
        puml_code: The PlantUML diagram code
        server_url: PlantUML server URL
        image_width: Optional width for the image (pixels)
        image_height: Optional height for the image (pixels)
        
    Returns:
        Path to the saved output file
    """
    template = Path(template_path)
    output = Path(output_path)
    
    if not template.exists():
        raise FileNotFoundError(f"Template not found at {template}")
    
    workbook = load_workbook(template)
    ws = workbook.active
    
    # Find marker cells
    puml_marker = _find_marker_cell(ws, PLANTUML_MARKER)
    image_marker = _find_marker_cell(ws, PLANTUML_IMAGE_MARKER)
    
    # Write the PlantUML code below the marker
    code_cell = ws.cell(row=puml_marker.row + 1, column=puml_marker.column)
    code_cell.value = puml_code
    
    # Make the cell wrap text for better display
    from openpyxl.styles import Alignment
    code_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Generate the image from PlantUML server
    image_data = generate_plantuml_image(puml_code, server_url=server_url)
    
    # Create an image object from the bytes
    img_stream = BytesIO(image_data)
    img = XLImage(img_stream)
    
    # Optionally resize the image
    if image_width:
        img.width = image_width
    if image_height:
        img.height = image_height
    
    # Place the image at the marker cell position
    # The anchor is the cell reference like "B5"
    from openpyxl.utils import get_column_letter
    anchor = f"{get_column_letter(image_marker.column)}{image_marker.row + 1}"
    ws.add_image(img, anchor)
    
    # Clear the marker text (optional - keep for reference)
    # image_marker.value = None
    
    # Save the workbook
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    
    return output


def write_plantuml_image_only(
    excel_path: str | Path,
    output_path: str | Path,
    puml_code: str,
    server_url: str = "http://localhost:8080",
    cell_anchor: str = "A1",
    sheet_name: str | None = None,
    image_width: int | None = None,
    image_height: int | None = None,
) -> Path:
    """Write a PlantUML image to a specific cell in an Excel file.
    
    This is a simpler version that doesn't require marker cells.
    
    Args:
        excel_path: Path to the source Excel file
        output_path: Path where the output file will be saved
        puml_code: The PlantUML diagram code
        server_url: PlantUML server URL
        cell_anchor: Cell reference where the image will be placed (e.g., "B5")
        sheet_name: Name of the sheet to write to (uses active sheet if None)
        image_width: Optional width for the image (pixels)
        image_height: Optional height for the image (pixels)
        
    Returns:
        Path to the saved output file
    """
    source = Path(excel_path)
    output = Path(output_path)
    
    if not source.exists():
        raise FileNotFoundError(f"Excel file not found at {source}")
    
    workbook = load_workbook(source)
    ws = workbook[sheet_name] if sheet_name else workbook.active
    
    # Generate the image from PlantUML server
    image_data = generate_plantuml_image(puml_code, server_url=server_url)
    
    # Create an image object from the bytes
    img_stream = BytesIO(image_data)
    img = XLImage(img_stream)
    
    # Optionally resize the image
    if image_width:
        img.width = image_width
    if image_height:
        img.height = image_height
    
    # Place the image at the specified cell
    ws.add_image(img, cell_anchor)
    
    # Save the workbook
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    
    return output
