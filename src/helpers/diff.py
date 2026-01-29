"""Diff helper for comparing code and writing to Excel."""

from __future__ import annotations

import difflib
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


@dataclass
class DiffBlock:
    """Represents a diff between old and new code."""
    old_code: str
    new_code: str
    old_label: str = "Old Code"
    new_label: str = "New Code"
    file_path: str | None = None


@dataclass
class DiffLine:
    """Represents a single line in a diff."""
    line_num_old: int | None
    line_num_new: int | None
    content: str
    change_type: Literal["unchanged", "added", "removed", "modified"]


def compute_diff(old_code: str, new_code: str) -> list[DiffLine]:
    """Compute line-by-line diff between old and new code.
    
    Args:
        old_code: Original code
        new_code: New/modified code
        
    Returns:
        List of DiffLine objects
    """
    old_lines = old_code.splitlines(keepends=True)
    new_lines = new_code.splitlines(keepends=True)
    
    diff_lines: list[DiffLine] = []
    
    # Use unified diff to get changes
    matcher = difflib.SequenceMatcher(None, old_lines, new_lines)
    
    old_line_num = 1
    new_line_num = 1
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                diff_lines.append(DiffLine(
                    line_num_old=old_line_num + k,
                    line_num_new=new_line_num + k,
                    content=old_lines[i1 + k].rstrip("\n"),
                    change_type="unchanged"
                ))
            old_line_num += (i2 - i1)
            new_line_num += (j2 - j1)
            
        elif tag == "delete":
            for k in range(i2 - i1):
                diff_lines.append(DiffLine(
                    line_num_old=old_line_num + k,
                    line_num_new=None,
                    content=old_lines[i1 + k].rstrip("\n"),
                    change_type="removed"
                ))
            old_line_num += (i2 - i1)
            
        elif tag == "insert":
            for k in range(j2 - j1):
                diff_lines.append(DiffLine(
                    line_num_old=None,
                    line_num_new=new_line_num + k,
                    content=new_lines[j1 + k].rstrip("\n"),
                    change_type="added"
                ))
            new_line_num += (j2 - j1)
            
        elif tag == "replace":
            # Show removed lines first, then added lines
            for k in range(i2 - i1):
                diff_lines.append(DiffLine(
                    line_num_old=old_line_num + k,
                    line_num_new=None,
                    content=old_lines[i1 + k].rstrip("\n"),
                    change_type="removed"
                ))
            old_line_num += (i2 - i1)
            
            for k in range(j2 - j1):
                diff_lines.append(DiffLine(
                    line_num_old=None,
                    line_num_new=new_line_num + k,
                    content=new_lines[j1 + k].rstrip("\n"),
                    change_type="added"
                ))
            new_line_num += (j2 - j1)
    
    return diff_lines


def write_diff_to_excel(
    old_code: str,
    new_code: str,
    output_path: str | Path,
    old_label: str = "Old Code",
    new_label: str = "New Code",
    file_path: str | None = None,
    side_by_side: bool = True,
) -> Path:
    """Write a code diff to Excel file.
    
    Args:
        old_code: Original code
        new_code: New/modified code
        output_path: Path for output Excel file
        old_label: Label for old code section
        new_label: Label for new code section
        file_path: Optional file path to show in header
        side_by_side: If True, show old and new side by side; else unified view
        
    Returns:
        Path to saved file
    """
    output = Path(output_path)
    wb = Workbook()
    
    # Colors
    HEADER_BG = "1E1F1C"
    ADDED_BG = "1E3A1E"      # Dark green
    REMOVED_BG = "3A1E1E"    # Dark red
    UNCHANGED_BG = "272822"  # Monokai dark
    
    ADDED_TEXT = "A6E22E"    # Green
    REMOVED_TEXT = "F92672"  # Pink/Red
    UNCHANGED_TEXT = "F8F8F2"  # White
    LINE_NUM_TEXT = "75715E"  # Gray
    
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    code_font = Font(name="Consolas", size=10)
    border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )
    
    if side_by_side:
        # =====================================================================
        # Side-by-Side View (2 sheets)
        # =====================================================================
        
        # Sheet 1: Old Code
        ws_old = wb.active
        ws_old.title = old_label
        _write_code_sheet(ws_old, old_code, old_label, file_path, 
                         UNCHANGED_BG, UNCHANGED_TEXT, LINE_NUM_TEXT, border)
        
        # Sheet 2: New Code
        ws_new = wb.create_sheet(title=new_label)
        _write_code_sheet(ws_new, new_code, new_label, file_path,
                         UNCHANGED_BG, UNCHANGED_TEXT, LINE_NUM_TEXT, border)
        
        # Sheet 3: Unified Diff
        ws_diff = wb.create_sheet(title="Diff View")
        _write_unified_diff_sheet(ws_diff, old_code, new_code, file_path,
                                  ADDED_BG, REMOVED_BG, UNCHANGED_BG,
                                  ADDED_TEXT, REMOVED_TEXT, UNCHANGED_TEXT,
                                  LINE_NUM_TEXT, border)
    else:
        # =====================================================================
        # Unified View Only
        # =====================================================================
        ws_diff = wb.active
        ws_diff.title = "Diff View"
        _write_unified_diff_sheet(ws_diff, old_code, new_code, file_path,
                                  ADDED_BG, REMOVED_BG, UNCHANGED_BG,
                                  ADDED_TEXT, REMOVED_TEXT, UNCHANGED_TEXT,
                                  LINE_NUM_TEXT, border)
    
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def _write_code_sheet(ws, code: str, label: str, file_path: str | None,
                      bg_color: str, text_color: str, line_num_color: str,
                      border: Border):
    """Write a single code block to a worksheet."""
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 120
    
    # Header
    header_fill = PatternFill(start_color="1E1F1C", end_color="1E1F1C", fill_type="solid")
    header_text = f"📄 {label}"
    if file_path:
        header_text += f" - {file_path}"
    
    header_cell = ws.cell(row=1, column=1, value=header_text)
    header_cell.font = Font(bold=True, size=11, color="FFFFFF")
    header_cell.fill = header_fill
    ws.merge_cells("A1:B1")
    
    # Column headers
    col_header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    for col, text in [(1, "Line"), (2, "Code")]:
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = col_header_fill
    
    # Code lines
    bg_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    lines = code.rstrip("\n").split("\n")
    
    for i, line in enumerate(lines, start=1):
        row = i + 2
        
        # Line number
        line_cell = ws.cell(row=row, column=1, value=i)
        line_cell.font = Font(name="Consolas", size=10, color=line_num_color)
        line_cell.fill = bg_fill
        line_cell.alignment = Alignment(horizontal="right")
        line_cell.border = border
        
        # Code
        code_cell = ws.cell(row=row, column=2, value=line)
        code_cell.font = Font(name="Consolas", size=10, color=text_color)
        code_cell.fill = bg_fill
        code_cell.alignment = Alignment(horizontal="left")
        code_cell.border = border


def _write_unified_diff_sheet(ws, old_code: str, new_code: str, file_path: str | None,
                               added_bg: str, removed_bg: str, unchanged_bg: str,
                               added_text: str, removed_text: str, unchanged_text: str,
                               line_num_color: str, border: Border):
    """Write unified diff view to worksheet."""
    ws.column_dimensions["A"].width = 8   # Old line
    ws.column_dimensions["B"].width = 8   # New line
    ws.column_dimensions["C"].width = 5   # Symbol
    ws.column_dimensions["D"].width = 120  # Code
    
    # Header
    header_fill = PatternFill(start_color="1E1F1C", end_color="1E1F1C", fill_type="solid")
    header_text = "📊 Diff View"
    if file_path:
        header_text += f" - {file_path}"
    
    header_cell = ws.cell(row=1, column=1, value=header_text)
    header_cell.font = Font(bold=True, size=11, color="FFFFFF")
    header_cell.fill = header_fill
    ws.merge_cells("A1:D1")
    
    # Legend
    ws.cell(row=2, column=1, value="Legend:").font = Font(bold=True, size=9, color="FFFFFF")
    ws.cell(row=2, column=1).fill = header_fill
    
    legend_items = [
        ("- Removed", removed_bg, removed_text),
        ("+ Added", added_bg, added_text),
        ("  Unchanged", unchanged_bg, unchanged_text),
    ]
    col = 2
    for text, bg, fg in legend_items:
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(size=9, color=fg)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        col += 1
    
    # Column headers
    col_header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    headers = [(1, "Old"), (2, "New"), (3, ""), (4, "Code")]
    for c, text in headers:
        cell = ws.cell(row=3, column=c, value=text)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = col_header_fill
    
    # Diff lines
    diff_lines = compute_diff(old_code, new_code)
    
    for i, diff_line in enumerate(diff_lines):
        row = i + 4
        
        if diff_line.change_type == "added":
            bg = added_bg
            text_color = added_text
            symbol = "+"
        elif diff_line.change_type == "removed":
            bg = removed_bg
            text_color = removed_text
            symbol = "-"
        else:
            bg = unchanged_bg
            text_color = unchanged_text
            symbol = ""
        
        bg_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        
        # Old line number
        old_cell = ws.cell(row=row, column=1, value=diff_line.line_num_old or "")
        old_cell.font = Font(name="Consolas", size=10, color=line_num_color)
        old_cell.fill = bg_fill
        old_cell.alignment = Alignment(horizontal="right")
        old_cell.border = border
        
        # New line number
        new_cell = ws.cell(row=row, column=2, value=diff_line.line_num_new or "")
        new_cell.font = Font(name="Consolas", size=10, color=line_num_color)
        new_cell.fill = bg_fill
        new_cell.alignment = Alignment(horizontal="right")
        new_cell.border = border
        
        # Symbol
        sym_cell = ws.cell(row=row, column=3, value=symbol)
        sym_cell.font = Font(name="Consolas", size=10, bold=True, color=text_color)
        sym_cell.fill = bg_fill
        sym_cell.alignment = Alignment(horizontal="center")
        sym_cell.border = border
        
        # Code
        code_cell = ws.cell(row=row, column=4, value=diff_line.content)
        code_cell.font = Font(name="Consolas", size=10, color=text_color)
        code_cell.fill = bg_fill
        code_cell.alignment = Alignment(horizontal="left")
        code_cell.border = border


def format_diff_text(old_code: str, new_code: str) -> str:
    """Format diff as text (unified diff format).
    
    Args:
        old_code: Original code
        new_code: New code
        
    Returns:
        Unified diff as string
    """
    old_lines = old_code.splitlines(keepends=True)
    new_lines = new_code.splitlines(keepends=True)
    
    diff = difflib.unified_diff(old_lines, new_lines, 
                                 fromfile="old", tofile="new",
                                 lineterm="")
    return "".join(diff)
