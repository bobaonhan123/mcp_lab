from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class MarkerNotFoundError(RuntimeError):
    """Raised when the expected marker cell is missing from the template."""


MARKERS = {
    "main": "marker[main_task]",
    "support": "marker[support_task]",
}


def _find_marker_cell(ws: Worksheet, marker: str) -> Cell:
    """Locate the cell whose value matches the marker string."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == marker:
                return cell
    raise MarkerNotFoundError(f"Could not find marker cell with value '{marker}'.")


def _apply_style(target: Cell, template_cell: Cell) -> None:
    """Copy cell style from template_cell onto target."""
    if template_cell._style:
        target._style = copy(template_cell._style)


def _clear_column(
    ws: Worksheet, column_index: int, start_row: int, end_row: int | None = None
) -> None:
    """Remove values in a column between start_row and end_row (inclusive)."""
    last_row = end_row if end_row is not None else ws.max_row
    for row in range(start_row, last_row + 1):
        ws.cell(row=row, column=column_index, value=None)


def _write_numbered_tasks(
    ws: Worksheet,
    column_index: int,
    start_row: int,
    tasks: Iterable[str],
    style_cell: Cell,
) -> int:
    """Write numbered tasks to the worksheet and return the next empty row index."""
    task_list = list(tasks)
    for idx, task in enumerate(task_list, start=1):
        row = start_row + idx - 1
        cell = ws.cell(row=row, column=column_index)
        _apply_style(cell, style_cell)
        cell.value = f"{idx}. {task.strip()}"
    return start_row + len(task_list)


def _ensure_capacity_for_main_tasks(
    ws: Worksheet, main_start_row: int, support_marker_row: int, task_count: int
) -> None:
    """Insert rows above the support section if main tasks need extra space."""
    available_rows = max(0, support_marker_row - main_start_row)
    overflow = max(0, task_count - available_rows)
    if overflow:
        ws.insert_rows(support_marker_row, amount=overflow)


def write_task_lists_to_excel(
    template_path: str | Path,
    output_path: str | Path,
    main_tasks: Iterable[str],
    support_tasks: Iterable[str],
) -> Path:
    """Fill the Excel template with the provided main and support tasks.

    The function respects marker cells in the workbook:
    - marker[main_task]: tasks are written beneath this cell
    - marker[support_task]: tasks are written beneath this cell

    The cell immediately below each marker acts as the formatting template and is
    applied to all generated rows.
    """
    template = Path(template_path)
    output = Path(output_path)

    if not template.exists():
        raise FileNotFoundError(f"Template not found at {template}")

    workbook = load_workbook(template)
    ws = workbook.active

    # Locate marker cells and formatting templates.
    main_marker = _find_marker_cell(ws, MARKERS["main"])
    support_marker = _find_marker_cell(ws, MARKERS["support"])
    main_style_cell = ws.cell(row=main_marker.row + 1, column=main_marker.column)
    support_style_cell = ws.cell(row=support_marker.row + 1, column=support_marker.column)

    # Make sure there is enough space for main tasks by shifting the support section down.
    main_tasks_list = list(main_tasks)
    _ensure_capacity_for_main_tasks(
        ws,
        main_marker.row + 1,
        support_marker.row,
        len(main_tasks_list),
    )

    # The support marker may have moved after insertion; re-find it and update template.
    support_marker = _find_marker_cell(ws, MARKERS["support"])
    support_style_cell = ws.cell(row=support_marker.row + 1, column=support_marker.column)

    # Clear existing values in both sections.
    main_start_row = main_marker.row + 1
    _clear_column(ws, main_marker.column, start_row=main_start_row, end_row=support_marker.row - 1)

    support_start_row = support_marker.row + 1
    _clear_column(ws, support_marker.column, start_row=support_start_row)

    # Write tasks with numbering; apply formatting from the template row.
    _write_numbered_tasks(
        ws,
        column_index=main_marker.column,
        start_row=main_start_row,
        tasks=main_tasks_list,
        style_cell=main_style_cell,
    )
    _write_numbered_tasks(
        ws,
        column_index=support_marker.column,
        start_row=support_start_row,
        tasks=list(support_tasks),
        style_cell=support_style_cell,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output

