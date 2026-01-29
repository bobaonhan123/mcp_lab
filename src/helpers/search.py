"""Search helper for searching in folders and writing results to Excel."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


@dataclass
class SearchMatch:
    """Represents a single search match."""
    file_path: str
    line_number: int
    line_content: str
    match_start: int
    match_end: int
    context_before: list[str] = field(default_factory=list)
    context_after: list[str] = field(default_factory=list)


@dataclass
class SearchResult:
    """Represents search results for a file."""
    file_path: str
    matches: list[SearchMatch] = field(default_factory=list)
    
    @property
    def match_count(self) -> int:
        return len(self.matches)


@dataclass
class SearchSummary:
    """Summary of all search results."""
    query: str
    is_regex: bool
    folder_path: str
    results: list[SearchResult] = field(default_factory=list)
    
    @property
    def total_matches(self) -> int:
        return sum(r.match_count for r in self.results)
    
    @property
    def files_with_matches(self) -> int:
        return len([r for r in self.results if r.match_count > 0])


# Common file extensions to search
DEFAULT_EXTENSIONS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".scala",
    ".html", ".css", ".scss", ".less", ".vue", ".svelte",
    ".json", ".yaml", ".yml", ".xml", ".toml", ".ini", ".cfg",
    ".md", ".txt", ".rst", ".sql", ".sh", ".bash", ".ps1", ".bat",
}

# Folders to exclude
DEFAULT_EXCLUDES = {
    "__pycache__", "node_modules", ".git", ".svn", ".hg",
    "venv", ".venv", "env", ".env",
    "dist", "build", "target", "out", "bin", "obj",
    ".idea", ".vscode", ".vs",
    "coverage", ".pytest_cache", ".mypy_cache",
}


def search_in_folder(
    folder_path: str | Path,
    query: str,
    is_regex: bool = False,
    case_sensitive: bool = False,
    include_extensions: set[str] | None = None,
    exclude_folders: set[str] | None = None,
    context_lines: int = 2,
    max_results: int = 1000,
) -> SearchSummary:
    """Search for a pattern in all files in a folder.
    
    Args:
        folder_path: Path to the folder to search
        query: Search query (string or regex pattern)
        is_regex: Whether query is a regex pattern
        case_sensitive: Whether search is case-sensitive
        include_extensions: File extensions to include (default: common code files)
        exclude_folders: Folder names to exclude (default: common non-code folders)
        context_lines: Number of context lines before/after match
        max_results: Maximum number of matches to return
        
    Returns:
        SearchSummary with all matches
    """
    folder = Path(folder_path)
    extensions = include_extensions or DEFAULT_EXTENSIONS
    excludes = exclude_folders or DEFAULT_EXCLUDES
    
    # Compile pattern
    flags = 0 if case_sensitive else re.IGNORECASE
    if is_regex:
        pattern = re.compile(query, flags)
    else:
        pattern = re.compile(re.escape(query), flags)
    
    summary = SearchSummary(
        query=query,
        is_regex=is_regex,
        folder_path=str(folder),
    )
    
    total_matches = 0
    
    for file_path in _iter_files(folder, extensions, excludes):
        if total_matches >= max_results:
            break
            
        try:
            result = _search_file(file_path, pattern, context_lines, max_results - total_matches)
            if result.match_count > 0:
                summary.results.append(result)
                total_matches += result.match_count
        except (UnicodeDecodeError, PermissionError, OSError):
            # Skip files that can't be read
            continue
    
    return summary


def _iter_files(folder: Path, extensions: set[str], excludes: set[str]) -> Iterator[Path]:
    """Iterate over files in folder, respecting extensions and excludes."""
    try:
        for item in folder.iterdir():
            if item.is_dir():
                if item.name not in excludes:
                    yield from _iter_files(item, extensions, excludes)
            elif item.is_file():
                if item.suffix.lower() in extensions:
                    yield item
    except PermissionError:
        pass


def _search_file(file_path: Path, pattern: re.Pattern, context_lines: int, max_matches: int) -> SearchResult:
    """Search for pattern in a single file."""
    result = SearchResult(file_path=str(file_path))
    
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()
    
    for i, line in enumerate(lines):
        if len(result.matches) >= max_matches:
            break
            
        match = pattern.search(line)
        if match:
            # Get context
            start_ctx = max(0, i - context_lines)
            end_ctx = min(len(lines), i + context_lines + 1)
            
            context_before = [lines[j].rstrip("\n") for j in range(start_ctx, i)]
            context_after = [lines[j].rstrip("\n") for j in range(i + 1, end_ctx)]
            
            result.matches.append(SearchMatch(
                file_path=str(file_path),
                line_number=i + 1,
                line_content=line.rstrip("\n"),
                match_start=match.start(),
                match_end=match.end(),
                context_before=context_before,
                context_after=context_after,
            ))
    
    return result


def format_search_results(summary: SearchSummary, show_context: bool = True) -> str:
    """Format search results as text.
    
    Args:
        summary: SearchSummary object
        show_context: Whether to show context lines
        
    Returns:
        Formatted text output
    """
    lines = []
    lines.append(f"🔍 Search Results for: '{summary.query}'")
    lines.append(f"   Folder: {summary.folder_path}")
    lines.append(f"   Matches: {summary.total_matches} in {summary.files_with_matches} files")
    lines.append("")
    
    for result in summary.results:
        lines.append(f"📄 {result.file_path} ({result.match_count} matches)")
        
        for match in result.matches:
            lines.append(f"   Line {match.line_number}: {match.line_content.strip()}")
            
            if show_context:
                for ctx_line in match.context_before:
                    lines.append(f"      | {ctx_line}")
                lines.append(f"   >> | {match.line_content}")
                for ctx_line in match.context_after:
                    lines.append(f"      | {ctx_line}")
                lines.append("")
        
        lines.append("")
    
    return "\n".join(lines)


def write_search_to_excel(
    summary: SearchSummary,
    output_path: str | Path,
    include_context: bool = True,
) -> Path:
    """Write search results to Excel file.
    
    Args:
        summary: SearchSummary object
        output_path: Path for output Excel file
        include_context: Whether to include context lines
        
    Returns:
        Path to saved file
    """
    output = Path(output_path)
    wb = Workbook()
    
    # Colors
    HEADER_BG = "1E1F1C"
    MATCH_BG = "3A3A1E"      # Dark yellow/olive
    CONTEXT_BG = "272822"    # Monokai dark
    SUMMARY_BG = "366092"    # Blue
    
    MATCH_TEXT = "E6DB74"    # Yellow
    CONTEXT_TEXT = "75715E"  # Gray
    NORMAL_TEXT = "F8F8F2"   # White
    
    border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )
    
    # =========================================================================
    # Sheet 1: Summary
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 80
    
    # Header
    summary_fill = PatternFill(start_color=SUMMARY_BG, end_color=SUMMARY_BG, fill_type="solid")
    ws_summary.cell(row=1, column=1, value="🔍 Search Results").font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary.cell(row=1, column=1).fill = summary_fill
    ws_summary.merge_cells("A1:B1")
    
    # Info
    info = [
        ("Query:", summary.query),
        ("Regex:", "Yes" if summary.is_regex else "No"),
        ("Folder:", summary.folder_path),
        ("Total Matches:", str(summary.total_matches)),
        ("Files with Matches:", str(summary.files_with_matches)),
    ]
    
    for i, (label, value) in enumerate(info, start=3):
        ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)
    
    # Files table
    row = len(info) + 5
    ws_summary.cell(row=row, column=1, value="Files:").font = Font(bold=True, size=12)
    row += 1
    
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    for col, text in [(1, "File"), (2, "Matches")]:
        cell = ws_summary.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    row += 1
    
    for result in summary.results:
        ws_summary.cell(row=row, column=1, value=result.file_path)
        ws_summary.cell(row=row, column=2, value=result.match_count)
        row += 1
    
    # =========================================================================
    # Sheet 2: All Matches
    # =========================================================================
    ws_matches = wb.create_sheet(title="All Matches")
    
    ws_matches.column_dimensions["A"].width = 50
    ws_matches.column_dimensions["B"].width = 8
    ws_matches.column_dimensions["C"].width = 100
    
    # Header
    match_header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    ws_matches.cell(row=1, column=1, value=f"🔍 Matches for: {summary.query}").font = Font(bold=True, size=11, color="FFFFFF")
    ws_matches.cell(row=1, column=1).fill = match_header_fill
    ws_matches.merge_cells("A1:C1")
    
    # Column headers
    col_header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    for col, text in [(1, "File"), (2, "Line"), (3, "Content")]:
        cell = ws_matches.cell(row=2, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = col_header_fill
    
    # Matches
    row = 3
    match_fill = PatternFill(start_color=MATCH_BG, end_color=MATCH_BG, fill_type="solid")
    context_fill = PatternFill(start_color=CONTEXT_BG, end_color=CONTEXT_BG, fill_type="solid")
    
    for result in summary.results:
        for match in result.matches:
            # Context before
            if include_context:
                for ctx_i, ctx_line in enumerate(match.context_before):
                    ctx_line_num = match.line_number - len(match.context_before) + ctx_i
                    ws_matches.cell(row=row, column=1, value="").fill = context_fill
                    
                    cell_line = ws_matches.cell(row=row, column=2, value=ctx_line_num)
                    cell_line.font = Font(name="Consolas", size=9, color=CONTEXT_TEXT)
                    cell_line.fill = context_fill
                    cell_line.border = border
                    
                    cell_code = ws_matches.cell(row=row, column=3, value=ctx_line)
                    cell_code.font = Font(name="Consolas", size=9, color=CONTEXT_TEXT)
                    cell_code.fill = context_fill
                    cell_code.border = border
                    row += 1
            
            # Match line
            cell_file = ws_matches.cell(row=row, column=1, value=match.file_path)
            cell_file.font = Font(size=9, color=NORMAL_TEXT)
            cell_file.fill = match_fill
            cell_file.border = border
            
            cell_line = ws_matches.cell(row=row, column=2, value=match.line_number)
            cell_line.font = Font(name="Consolas", size=10, bold=True, color=MATCH_TEXT)
            cell_line.fill = match_fill
            cell_line.alignment = Alignment(horizontal="right")
            cell_line.border = border
            
            cell_code = ws_matches.cell(row=row, column=3, value=match.line_content)
            cell_code.font = Font(name="Consolas", size=10, color=MATCH_TEXT)
            cell_code.fill = match_fill
            cell_code.border = border
            row += 1
            
            # Context after
            if include_context:
                for ctx_i, ctx_line in enumerate(match.context_after):
                    ctx_line_num = match.line_number + ctx_i + 1
                    ws_matches.cell(row=row, column=1, value="").fill = context_fill
                    
                    cell_line = ws_matches.cell(row=row, column=2, value=ctx_line_num)
                    cell_line.font = Font(name="Consolas", size=9, color=CONTEXT_TEXT)
                    cell_line.fill = context_fill
                    cell_line.border = border
                    
                    cell_code = ws_matches.cell(row=row, column=3, value=ctx_line)
                    cell_code.font = Font(name="Consolas", size=9, color=CONTEXT_TEXT)
                    cell_code.fill = context_fill
                    cell_code.border = border
                    row += 1
                
                # Blank row between matches
                row += 1
    
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output
